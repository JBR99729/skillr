#!/usr/bin/env python3
"""Build SkillrHub curriculum manifests and SEO topic guide pages.

The workbook remains the source of truth. This script intentionally keeps
interactive quiz generation separate from SEO topic-page generation so complex
question types can later be routed to either browser interactions or printable
PDF/pen-and-paper tasks without rewriting the page structure.
"""

from __future__ import annotations

import argparse
import html
import json
import re
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


SITE_ORIGIN = "https://skillrhub.com"
GA_ID = "G-8P22BET45N"
ADSENSE_CLIENT = "ca-pub-7734963540104771"
SUPPORT_EMAIL = "skillrhublearning@gmail.com"

LEARNING_AREAS = ("Mathematics", "Science", "English")
LEARNING_AREA_ORDER = {name: index for index, name in enumerate(LEARNING_AREAS)}
LEVELS = ["Foundation Year", *[f"Year {number}" for number in range(1, 11)]]

SUBJECT_SLUGS = {
    "English": "english",
    "Mathematics": "maths",
    "Science": "science",
}

QUIZ_SUBJECT_SLUGS = {
    "English": "english",
    "Mathematics": "math",
    "Science": "science",
}

FIRST_NATIONS_RE = re.compile(
    r"\b(First Nations?|Aboriginal|Torres Strait Islander|Indigenous)\b",
    re.IGNORECASE,
)

CODE_RE = re.compile(r"^AC9[EMS][A-Z0-9]+$")
ELABORATION_RE = re.compile(r"^(AC9[EMS][A-Z0-9]+)_E(\d+)$")


TITLE_OVERRIDES = {
    "AC9M10P01": "Conditional probability language: if-then, given, of and knowing that",
    "AC9M1N01": "Numbers to 120",
    "AC9M1N02": "Partitioning tens and ones",
    "AC9M1N03": "Skip counting and equal groups",
    "AC9M1N04": "Addition and subtraction within 20",
    "AC9M1N05": "Additive problem solving and money",
    "AC9M1N06": "Equal sharing and grouping",
    "AC9M1A01": "Skip-counting pattern sequences",
    "AC9M1A02": "Repeating patterns",
    "AC9M1M01": "Comparing length, mass, capacity and duration",
    "AC9M1M02": "Measuring length with informal units",
    "AC9M1M03": "Time: years, months, weeks, days and hours",
    "AC9M1SP01": "Familiar shapes and objects",
    "AC9M1SP02": "Directions and location",
    "AC9M1ST01": "Collecting categorical data",
    "AC9M1ST02": "Representing and comparing data",
    "AC9S1U01": "Needs of plants and animals",
    "AC9S1U02": "Daily and seasonal changes",
    "AC9S1U03": "Pushes and pulls",
    "AC9S1H01": "Science in daily life",
    "AC9S1I01": "Questions, patterns and predictions",
    "AC9S1I02": "Safe investigation procedures",
    "AC9S1I03": "Making and recording observations",
    "AC9S1I04": "Sorting data and representing patterns",
    "AC9S1I05": "Comparing observations and predictions",
    "AC9S1I06": "Communicating scientific ideas",
}


@dataclass
class BuildContext:
    repo_root: Path
    workbook: Path
    manifest_path: Path
    units: list[dict[str, Any]]


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value)
    text = re.sub(r"<[^>]+>", " ", text)
    text = html.unescape(text)
    text = text.replace("\\(", "").replace("\\)", "")
    text = text.replace("\u00a0", " ")
    text = text.replace("–", "-")
    text = re.sub(r"\s+", " ", text).strip()
    return text


def slugify(value: str, max_length: int = 64) -> str:
    value = value.lower()
    value = value.replace("&", " and ")
    value = re.sub(r"[^a-z0-9]+", "-", value)
    value = re.sub(r"-+", "-", value).strip("-")
    if len(value) > max_length:
        value = value[:max_length].rsplit("-", 1)[0] or value[:max_length]
    return value or "topic"


def level_number(level: str) -> int:
    if level == "Foundation Year":
        return 0
    match = re.search(r"\d+", level)
    if not match:
        raise ValueError(f"Unsupported level: {level}")
    return int(match.group(0))


def level_label(level: str) -> str:
    return "Foundation" if level == "Foundation Year" else level


def year_folder(level: str) -> str:
    number = level_number(level)
    return "foundation" if number == 0 else f"year{number}"


def quiz_year_segment(level: str) -> str:
    number = level_number(level)
    return "grade-k" if number == 0 else f"year-{number}"


def nsw_stage(level: str) -> str:
    number = level_number(level)
    if number == 0:
        return "Early Stage 1"
    if number <= 2:
        return "Stage 1"
    if number <= 4:
        return "Stage 2"
    if number <= 6:
        return "Stage 3"
    if number <= 8:
        return "Stage 4"
    return "Stage 5"


def nz_level(level: str) -> str:
    number = level_number(level)
    if number <= 2:
        return "Level 1"
    if number <= 4:
        return "Level 2"
    if number <= 6:
        return "Level 3"
    if number <= 8:
        return "Level 4"
    return "Level 5"


def england_key_stage(level: str) -> str:
    number = level_number(level)
    if number <= 2:
        return "Key Stage 1"
    if number <= 6:
        return "Key Stage 2"
    if number <= 9:
        return "Key Stage 3"
    return "Key Stage 4"


def class_label(level: str) -> str:
    number = level_number(level)
    return "Kindergarten / Foundation" if number == 0 else f"Class {number}"


def legacy_title_from_description(code: str, description: str) -> str:
    """Return the historic short title used to keep existing URLs stable."""
    if code in TITLE_OVERRIDES:
        return TITLE_OVERRIDES[code]

    description = clean_text(description)
    description = re.sub(
        r"^(understand|explore|recognise|compare|describe|discuss|listen to and discuss|use|create|read|write|spell|segment|orally|pose|suggest|make|sort|identify|quantify)\s+",
        "",
        description,
        flags=re.IGNORECASE,
    )
    words = description.split()
    title = " ".join(words[:10])
    title = title.rstrip(" ,;:.")
    if len(words) > 10:
        title += "..."
    return title[:1].upper() + title[1:]


def title_from_description(code: str, description: str) -> str:
    """Return a complete, readable display title without literal truncation."""
    if code in TITLE_OVERRIDES:
        return TITLE_OVERRIDES[code]

    # Foundation Maths is maintained by its dedicated quality build and must
    # keep its existing display titles until that build is intentionally run.
    if code.startswith("AC9MF"):
        return legacy_title_from_description(code, description)

    description = clean_text(description).rstrip(" .")
    return description[:1].upper() + description[1:]


def source_links() -> list[tuple[str, str]]:
    return [
        ("Australian Curriculum Version 9.0", "https://www.australiancurriculum.edu.au/"),
        ("Victorian Curriculum F-10", "https://f10.vcaa.vic.edu.au/"),
        ("NSW Curriculum", "https://curriculum.nsw.edu.au/"),
        ("Common Core State Standards", "https://corestandards.org/"),
        ("Next Generation Science Standards", "https://www.nextgenscience.org/"),
        ("England National Curriculum", "https://www.gov.uk/government/collections/national-curriculum"),
        ("New Zealand Curriculum", "https://newzealandcurriculum.tahurangi.education.govt.nz/"),
        ("NCERT textbooks", "https://ncert.nic.in/textbook.php"),
    ]


def unit_url(unit: dict[str, Any]) -> str:
    return f"/{unit['yearFolder']}/{unit['subjectSlug']}/{unit['unitSlug']}/"


def quiz_url(unit: dict[str, Any], mode: str) -> str:
    return f"/quiz/{unit['quizYearSegment']}/{unit['quizSubjectSlug']}/{unit['code'].lower()}/{mode}/"


def worksheet_url(unit: dict[str, Any]) -> str:
    return f"/quiz/{unit['quizYearSegment']}/{unit['quizSubjectSlug']}/{unit['code'].lower()}/worksheet/"


def teacher_slide_url(unit: dict[str, Any]) -> str:
    if unit["code"] in {"AC9M3M03", "AC9M3M04", "AC9S3U04"}:
        return f"/{unit['yearFolder']}/{unit['subjectSlug']}/{unit['unitSlug']}/teacher-deck/"
    if unit["yearFolder"] in {"foundation", "year1"} or (
        unit["yearFolder"] == "year3" and unit["subjectSlug"] == "maths"
    ):
        return f"/worksheets/{unit['yearFolder']}/{unit['subjectSlug']}/teacher-slides/live.html?code={unit['code']}"
    return f"/worksheets/{unit['yearFolder']}/{unit['subjectSlug']}/teacher-slides/{unit['code'].lower()}-teacher-slide.pdf"


def build_manifest(workbook: Path, repo_root: Path) -> list[dict[str, Any]]:
    wb = load_workbook(workbook, read_only=True, data_only=True)
    ws = wb["Learning areas"]
    headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
    index = {header: position for position, header in enumerate(headers) if header}

    units_by_code: dict[str, dict[str, Any]] = {}
    pending_elaborations: dict[str, list[dict[str, Any]]] = defaultdict(list)
    level_descriptions: dict[tuple[str, str], str] = {}
    current_strand: dict[tuple[str, str], str] = {}
    current_substrand: dict[tuple[str, str], str] = {}
    source_order = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        learning_area = clean_text(row[index["Learning Area"]])
        level = clean_text(row[index["Level"]])
        code = clean_text(row[index["Code"]])

        if learning_area not in LEARNING_AREAS or level not in LEVELS:
            continue

        key = (learning_area, level)

        level_description = clean_text(row[index["Level Description"]])
        if level_description:
            level_descriptions[key] = level_description

        strand = clean_text(row[index["Strand"]])
        if strand:
            current_strand[key] = strand
            current_substrand[key] = ""

        sub_strand = clean_text(row[index["Sub-Strand"]])
        if sub_strand:
            current_substrand[key] = sub_strand

        content_description = clean_text(row[index["Content Description"]])
        elaboration_text = clean_text(row[index["Elaboration"]])

        if CODE_RE.match(code) and content_description:
            source_order += 1
            title = title_from_description(code, content_description)
            legacy_title = legacy_title_from_description(code, content_description)
            subject_slug = SUBJECT_SLUGS[learning_area]
            unit = {
                "learningArea": learning_area,
                "subject": learning_area,
                "subjectSlug": subject_slug,
                "quizSubjectSlug": QUIZ_SUBJECT_SLUGS[learning_area],
                "level": level,
                "levelLabel": level_label(level),
                "yearNumber": level_number(level),
                "yearFolder": year_folder(level),
                "quizYearSegment": quiz_year_segment(level),
                "code": code,
                "title": title,
                "unitSlug": f"{code.lower()}-{slugify(legacy_title)}",
                "sourceOrder": source_order,
                "strand": current_strand.get(key, ""),
                "subStrand": current_substrand.get(key, ""),
                "description": content_description,
                "levelDescription": level_descriptions.get(key, ""),
                "questionEligible": not FIRST_NATIONS_RE.search(content_description),
                "elaborations": [],
            }
            units_by_code[code] = unit
            if code in pending_elaborations:
                unit["elaborations"].extend(pending_elaborations.pop(code))
            continue

        match = ELABORATION_RE.match(code)
        if match and elaboration_text:
            base_code, number = match.groups()
            elaboration = {
                "code": code,
                "number": f"E{number}",
                "text": elaboration_text,
                "questionEligible": not FIRST_NATIONS_RE.search(elaboration_text),
            }
            if base_code in units_by_code:
                units_by_code[base_code]["elaborations"].append(elaboration)
            else:
                pending_elaborations[base_code].append(elaboration)

    units = list(units_by_code.values())
    units.sort(
        key=lambda item: (
            item["yearNumber"],
            LEARNING_AREA_ORDER[item["learningArea"]],
            item["sourceOrder"],
        )
    )

    for unit in units:
        unit["url"] = unit_url(unit)
        unit["practiceUrl"] = quiz_url(unit, "practice")
        unit["testUrl"] = quiz_url(unit, "test")
        unit["worksheetUrl"] = worksheet_url(unit)
        unit["teacherSlideUrl"] = teacher_slide_url(unit)
        unit["questionCoverage"] = [
            {"code": unit["code"], "label": "Content description", "text": unit["description"], "questionEligible": unit["questionEligible"]},
            *unit["elaborations"],
        ]

    return units


def write_json(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")


def h(value: Any) -> str:
    return html.escape(clean_text(value), quote=True)


def paragraph_from_list(items: list[str], fallback: str) -> str:
    clean_items = [clean_text(item) for item in items if clean_text(item)]
    if not clean_items:
        return fallback
    return " ".join(clean_items)


def page_head(title: str, description: str, canonical_path: str, schema: dict[str, Any]) -> str:
    schema_json = json.dumps(schema, ensure_ascii=False)
    return f"""<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <meta name="google-adsense-account" content="{ADSENSE_CLIENT}">
  <title>{h(title)}</title>
  <meta name="description" content="{h(description)}">
  <meta name="robots" content="index, follow">
  <link rel="canonical" href="{SITE_ORIGIN}{canonical_path}">
  <meta property="og:title" content="{h(title)}">
  <meta property="og:description" content="{h(description)}">
  <meta property="og:type" content="website">
  <meta property="og:url" content="{SITE_ORIGIN}{canonical_path}">
  <meta property="og:site_name" content="SkillrHub">
  <meta name="twitter:card" content="summary">
  <meta name="theme-color" content="#1a3a72">
  <link rel="manifest" href="/manifest.webmanifest">
  <link rel="apple-touch-icon" href="/icons/apple-touch-icon.png">
  <link rel="stylesheet" href="/style.css">
  <link rel="stylesheet" href="/assets/curriculum.css?v=3">
  <script async src="https://www.googletagmanager.com/gtag/js?id={GA_ID}"></script>
  <script>
    window.dataLayer = window.dataLayer || [];
    function gtag(){{dataLayer.push(arguments);}}
    gtag("js", new Date());
    gtag("config", "{GA_ID}");
  </script>
  <script async src="https://pagead2.googlesyndication.com/pagead/js/adsbygoogle.js?client={ADSENSE_CLIENT}" crossorigin="anonymous"></script>
  <script type="application/ld+json">{schema_json}</script>
</head>"""


def nav_html(current: str) -> str:
    return f"""<nav class="main-nav">
  <a href="/">Home</a>
  <a href="/sitemap.html">Sitemap</a>
  <a href="/about.html">About</a>
  <a href="/contact.html">Contact</a>
  <a href="/how-to-use-skillr.html">How to use Skillr</a>
  <a href="/privacy-policy.html">Privacy Policy</a>
</nav>
<nav aria-label="Breadcrumb" class="breadcrumb">
  <ol>
    <li><a href="/">Home</a></li>
    <li aria-current="page">{h(current)}</li>
  </ol>
</nav>"""


def ad_slot(kind: str = "inline") -> str:
    return f"""<aside class="curriculum-ad-slot {kind}" aria-label="Sponsored content">
  <p class="curriculum-ad-label">Sponsored</p>
  <ins class="adsbygoogle"
       style="display:block"
       data-ad-client="{ADSENSE_CLIENT}"
       data-ad-format="auto"
       data-full-width-responsive="true"></ins>
  <script>(adsbygoogle = window.adsbygoogle || []).push({{}});</script>
</aside>"""


def year_navigation_table(current_level: str) -> str:
    rows = []
    for level in LEVELS:
        folder = year_folder(level)
        label = level_label(level)
        status = "Current pilot" if level == current_level else "Planned / existing hub"
        rows.append(
            f"<tr><td><a href='/{folder}/'>{h(label)}</a></td><td>{h(status)}</td><td>Maths, Science, English curriculum units</td></tr>"
        )
    return f"""<div class="curriculum-table-wrap">
  <table class="year-nav-table">
    <thead><tr><th>Year</th><th>Status</th><th>Subjects</th></tr></thead>
    <tbody>{''.join(rows)}</tbody>
  </table>
</div>"""


def unit_card(unit: dict[str, Any]) -> str:
    elaboration_count = len(unit["elaborations"])
    eligible_count = sum(1 for item in unit["questionCoverage"] if item.get("questionEligible"))
    return f"""<article class="curriculum-unit-card" id="{h(unit['code'].lower())}">
  <span class="curriculum-badge">{h(unit['code'])}</span>
  <h3>{h(unit['description'])}</h3>
  <p>{h(unit['description'])}</p>
  <div class="unit-meta">
    <span class="curriculum-chip">{h(unit['strand'] or unit['subject'])}</span>
    <span class="curriculum-chip">{elaboration_count} elaborations</span>
    <span class="curriculum-chip">{eligible_count} question coverage points</span>
  </div>
  <div class="unit-action-row">
    <a class="primary" href="{h(unit['url'])}">Read topic guide</a>
    <a href="{h(unit['url'])}#teacher-slide">Teacher slide</a>
    <a href="{h(unit['url'])}#worksheet">Worksheets</a>
    <a href="{h(unit['url'])}#practice">Practice</a>
    <a href="{h(unit['url'])}#test">Test</a>
  </div>
</article>"""


def generate_landing_page(ctx: BuildContext, level: str) -> None:
    units = [unit for unit in ctx.units if unit["level"] == level]
    by_subject = {subject: [unit for unit in units if unit["learningArea"] == subject] for subject in LEARNING_AREAS}
    label = level_label(level)
    landing_path = f"/{year_folder(level)}/curriculum/"
    title = f"{label} Curriculum Units | Maths, Science and English Worksheets"
    description = (
        f"{label} curriculum units for Maths, Science and English with topic guides, "
        "worksheets, practice and tests aligned to Australian Curriculum Version 9.0."
    )
    subject_tiles = []
    for subject, subject_units in by_subject.items():
        subject_tiles.append(
            f"""<a class="subject-tile" href="#{slugify(subject)}">
  <span class="curriculum-badge">{len(subject_units)} units</span>
  <h2>{h(subject if subject != 'Mathematics' else 'Maths')}</h2>
  <p>{h(label)} {h(subject)} units arranged by curriculum code. Open the guide, worksheet, practice or test for each unit.</p>
</a>"""
        )

    sections = []
    for subject, subject_units in by_subject.items():
        subject_label = "Maths" if subject == "Mathematics" else subject
        sections.append(
            f"""<section class="curriculum-panel" id="{slugify(subject)}">
  <p class="curriculum-eyebrow">{h(label)} {h(subject_label)}</p>
  <h2>{h(subject_label)} unit matrix</h2>
  <p>Each unit is built from the workbook curriculum code and includes topic guide, worksheet, practice and test entry points.</p>
  <div class="unit-matrix">{''.join(unit_card(unit) for unit in subject_units)}</div>
</section>"""
        )

    schema = {
        "@context": "https://schema.org",
        "@type": "CollectionPage",
        "name": title,
        "url": f"{SITE_ORIGIN}{landing_path}",
        "description": description,
        "isPartOf": {"@type": "WebSite", "name": "SkillrHub", "url": SITE_ORIGIN},
    }

    html_text = f"""<!DOCTYPE html>
<html lang="en-AU">
{page_head(title, description, landing_path, schema)}
<body class="curriculum-shell">
<div class="curriculum-page">
  {nav_html(label + " curriculum")}
  <header class="curriculum-hero">
    <p class="curriculum-eyebrow">SkillrHub curriculum pilot</p>
    <h1>{h(label)} curriculum units for Maths, Science and English</h1>
    <p class="curriculum-hero__lead">This page is the new curriculum-code landing pattern: one year page, three subject tiles, and unit cards with topic guide, worksheet, practice and test actions.</p>
    <div class="curriculum-actions">
      <a class="curriculum-button primary" href="#mathematics">Maths units</a>
      <a class="curriculum-button" href="#science">Science units</a>
      <a class="curriculum-button" href="#english">English units</a>
    </div>
  </header>

  <section class="curriculum-panel">
    <h2>Navigate by year</h2>
    <p>SkillrHub is being built year by year and unit by unit. This table keeps every year internally linked for learners and search engines.</p>
    {year_navigation_table(level)}
  </section>

  {ad_slot("inline")}

  <section class="subject-tile-grid" aria-label="{h(label)} subject tiles">
    {''.join(subject_tiles)}
  </section>

  {''.join(sections)}

  <section class="curriculum-panel">
    <h2>Quality and maintenance standard</h2>
    <p>Practice banks are planned as 56 questions per curriculum code, test banks as 24 questions per curriculum code, with only 8 questions pulled per attempt. Printable worksheets stay clean with 10 unique printable questions. Complex multi-step work can be routed to downloadable PDF notes or worksheets instead of forcing complex browser logic.</p>
  </section>

  <p class="curriculum-footer-meta">Last reviewed: 10 August 2026. Feedback: <a href="mailto:{SUPPORT_EMAIL}">{SUPPORT_EMAIL}</a>.</p>
</div>
<script>
  window.skillrPageMeta = {{
    pageType: "year curriculum landing",
    year: "{h(label)}",
    supportEmail: "{SUPPORT_EMAIL}"
  }};
</script>
<script src="/assets/access.js?v=1"></script>
<script src="/assets/report-issue.js?v=1"></script>
<script src="/pwa-register.js"></script>
</body>
</html>
"""
    output = ctx.repo_root / year_folder(level) / "curriculum" / "index.html"
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(html_text, encoding="utf-8")


def mapping_rows(unit: dict[str, Any]) -> list[tuple[str, str, str]]:
    level = unit["level"]
    label = unit["levelLabel"]
    subject = "Maths" if unit["learningArea"] == "Mathematics" else unit["learningArea"]
    strand = unit["subStrand"] or unit["strand"] or subject
    grade = "Kindergarten" if unit["yearNumber"] == 0 else f"Grade {unit['yearNumber']}"
    return [
        ("Australia", "Australian Curriculum v9.0", f"{unit['code']} — {unit['description']}"),
        ("Victoria", "Victorian Curriculum F-10", f"{label} {subject}: closest match in {strand}. Use this page as a VIC-aligned practice and worksheet reference."),
        ("NSW", "NSW Curriculum", f"{nsw_stage(level)} {subject}: closest content focus for {unit['title']} and related outcomes."),
        ("United States", "Common Core / NGSS", f"{grade} {'Common Core Mathematics/ELA' if subject != 'Science' else 'NGSS science'} closest topic match for {unit['title']}."),
        ("England / UK", "National Curriculum", f"{england_key_stage(level)} / Year {max(1, unit['yearNumber'])}: closest programme-of-study match for {unit['title']}."),
        ("Canada", "Provincial and territory curricula", f"{grade} closest topic match. Canada varies by province, so use this as a broad Ontario/BC-style learning outcome reference."),
        ("New Zealand", "New Zealand Curriculum", f"{nz_level(level)} {subject}: closest achievement-objective topic for {unit['title']}."),
        ("India", "NCERT / CBSE", f"{class_label(level)} closest NCERT/CBSE topic match for {unit['title']}."),
    ]


def topic_keywords(unit: dict[str, Any]) -> list[str]:
    label = unit["levelLabel"]
    subject = "Maths" if unit["learningArea"] == "Mathematics" else unit["learningArea"]
    code = unit["code"]
    title = unit["title"]
    phrases = [
        f"{label} {subject} {title}",
        f"{code} worksheet",
        f"{code} practice questions",
        f"{code} test",
        f"Australian Curriculum {code}",
        f"{label} {subject} worksheets Australia",
        f"free printable {label} {subject} worksheets",
        f"{label} {subject} topic guide",
        f"Victorian Curriculum {label} {subject} {title}",
        f"NSW {nsw_stage(unit['level'])} {subject} {title}",
        f"UK {england_key_stage(unit['level'])} {subject} {title}",
        f"NZ Curriculum {nz_level(unit['level'])} {subject} {title}",
        f"CBSE NCERT {class_label(unit['level'])} {subject} {title}",
    ]
    return phrases


def common_mistakes(unit: dict[str, Any]) -> list[str]:
    subject = unit["learningArea"]
    if subject == "Mathematics":
        return [
            "Rushing to a rule before checking the concrete model, drawing or number sentence.",
            "Using the correct answer once but not being able to explain why it works.",
            "Mixing up similar vocabulary such as more/less, before/after, longer/shorter or equal groups/sharing.",
        ]
    if subject == "Science":
        return [
            "Describing what is seen without connecting it to a question, pattern or prediction.",
            "Using everyday words only and missing simple scientific vocabulary.",
            "Changing more than one thing in an investigation, which makes the comparison unfair.",
        ]
    return [
        "Answering from memory without using clues from the text, sentence or image.",
        "Knowing the idea orally but not transferring it into reading, writing or speaking.",
        "Missing punctuation, word order or vocabulary clues that change meaning.",
    ]


def learning_steps(unit: dict[str, Any]) -> list[str]:
    description = unit["description"]
    eligible_elaborations = [
        item["text"] for item in unit["elaborations"] if item.get("questionEligible")
    ]
    return [
        f"Start with the main idea: {description}.",
        "Use concrete examples, pictures, oral explanation and short written responses before moving to independent practice.",
        paragraph_from_list(
            eligible_elaborations[:2],
            "Practise with varied examples so students can recognise the concept in more than one form.",
        ),
        "Finish with a mixed activity so students must choose the correct strategy rather than copy the last example.",
    ]


def generate_topic_page(ctx: BuildContext, unit: dict[str, Any], all_year_units: list[dict[str, Any]]) -> None:
    label = unit["levelLabel"]
    subject_label = "Maths" if unit["learningArea"] == "Mathematics" else unit["learningArea"]
    canonical_path = unit["url"]
    title = f"{unit['code']} {unit['title']} | {label} {subject_label} Topic Guide"
    description = (
        f"{unit['code']} {label} {subject_label} topic guide for {unit['title']}, with curriculum mapping, "
        "keywords, worksheets, practice and test links."
    )
    related_same_subject = [
        item for item in all_year_units if item["learningArea"] == unit["learningArea"] and item["code"] != unit["code"]
    ]
    current_index = [item["code"] for item in all_year_units if item["learningArea"] == unit["learningArea"]].index(unit["code"])
    subject_units = [item for item in all_year_units if item["learningArea"] == unit["learningArea"]]
    previous_unit = subject_units[current_index - 1] if current_index > 0 else None
    next_unit = subject_units[current_index + 1] if current_index + 1 < len(subject_units) else None

    keyword_items = "".join(f"<li>{h(keyword)}</li>" for keyword in topic_keywords(unit))
    coverage_items = "".join(
        f"<li><strong>{h(item.get('number') or item.get('label') or item.get('code'))}:</strong> {h(item['text'])}{'' if item.get('questionEligible') else ' <em>(teaching context)</em>'}</li>"
        for item in unit["questionCoverage"]
    )
    steps = "".join(f"<li>{h(step)}</li>" for step in learning_steps(unit))
    mistakes = "".join(f"<li>{h(item)}</li>" for item in common_mistakes(unit))
    mapping = "".join(
        f"<tr><td>{h(country)}</td><td>{h(system)}</td><td>{h(match)}</td></tr>"
        for country, system, match in mapping_rows(unit)
    )
    source_items = "".join(
        f"<li><a href='{h(url)}' rel='nofollow noopener' target='_blank'>{h(name)}</a></li>"
        for name, url in source_links()
    )
    related_items = "".join(
        f"<li><a href='{h(item['url'])}'>{h(item['code'])}: {h(item['title'])}</a></li>"
        for item in related_same_subject[:8]
    )
    action_links = [
        '<a class="primary" href="#topic-guide">Topic guide</a>',
        f'<a href="{h(unit["worksheetUrl"])}">Worksheet</a>',
        f'<a href="{h(unit["practiceUrl"])}">Practice</a>',
        f'<a href="{h(unit["testUrl"])}">Test</a>',
    ]
    topic_actions = "".join(action_links)

    vertical_links = []
    if unit["yearNumber"] > 0:
        prev_folder = "foundation" if unit["yearNumber"] == 1 else f"year{unit['yearNumber'] - 1}"
        vertical_links.append(f"<a href='/{prev_folder}/'>Previous year</a>")
    if unit["yearNumber"] < 10:
        vertical_links.append(f"<a href='/year{unit['yearNumber'] + 1}/'>Next year</a>")
    if previous_unit:
        vertical_links.append(f"<a href='{h(previous_unit['url'])}'>Previous {h(subject_label)} unit</a>")
    if next_unit:
        vertical_links.append(f"<a href='{h(next_unit['url'])}'>Next {h(subject_label)} unit</a>")

    schema = {
        "@context": "https://schema.org",
        "@type": "LearningResource",
        "name": title,
        "description": description,
        "url": f"{SITE_ORIGIN}{canonical_path}",
        "educationalLevel": label,
        "learningResourceType": "Topic guide",
        "teaches": unit["description"],
        "isPartOf": {"@type": "WebSite", "name": "SkillrHub", "url": SITE_ORIGIN},
    }

    html_text = f"""<!DOCTYPE html>
<html lang="en-AU">
{page_head(title, description, canonical_path, schema)}
<body class="curriculum-shell">
<div class="curriculum-page">
  {nav_html(unit['code'])}
  <header class="curriculum-hero">
    <p class="curriculum-eyebrow">{h(unit['code'])} • {h(label)} {h(subject_label)}</p>
    <h1>{h(unit['code'])}: {h(unit['title'])}</h1>
    <p class="curriculum-hero__lead">{h(unit['description'])}</p>
    <div class="topic-action-row">
      {topic_actions}
    </div>
    <button class="report-issue-button" type="button" data-report-issue>Report issue</button>
  </header>

  <main class="curriculum-layout">
    <div>
      <section class="curriculum-topic-section" id="topic-guide">
        <h2>What students learn in {h(unit['code'])}</h2>
        <p>This unit helps students build a clear, usable understanding of <strong>{h(unit['title'])}</strong>. The goal is not to memorise one answer pattern. Students should be able to explain the idea, recognise it in a new example and apply it in a short practice or worksheet task.</p>
        <ul class="curriculum-check-list">{steps}</ul>
      </section>

      <section class="curriculum-topic-section">
        <h2>Curriculum coverage and elaborations</h2>
        <p>The content description and elaborations below show the curriculum ideas taught in this unit. Items marked as teaching context support lesson planning.</p>
        <ul>{coverage_items}</ul>
      </section>

      <section class="curriculum-topic-section">
        <h2>How to use this unit</h2>
        <p>Read the topic guide and use the teacher slide for instruction. Students can then use the worksheet for written work, open Practice for supported feedback, or take the Test when they are ready.</p>
      </section>

      <section class="curriculum-topic-section teacher-resource" id="teacher-slide">
        <p class="curriculum-eyebrow">Teacher resource</p>
        <h2>{h(unit['code'])} teacher slide</h2>
        <p>Use this one-page PDF to introduce the key idea, vocabulary and teaching sequence before students begin the activities.</p>
        <a class="curriculum-button primary" href="{h(unit['teacherSlideUrl'])}" target="_blank" rel="noopener">Open teacher slide (PDF)</a>
      </section>

      <section class="curriculum-topic-section">
        <h2>Common mistakes to watch for</h2>
        <ul>{mistakes}</ul>
      </section>

      <section class="curriculum-topic-section">
        <h2>International curriculum mapping</h2>
        <p>This table gives closest-topic mapping for search and planning. The Australian Curriculum code is exact; overseas entries are broad equivalents because each jurisdiction structures outcomes differently.</p>
        <div class="curriculum-table-wrap">
          <table class="curriculum-map-table">
            <thead><tr><th>Region</th><th>Curriculum</th><th>Closest mapping</th></tr></thead>
            <tbody>{mapping}</tbody>
          </table>
        </div>
      </section>

      <section class="curriculum-topic-section">
        <h2>Related {h(label)} {h(subject_label)} topics</h2>
        <ul class="curriculum-related-list">{related_items}</ul>
      </section>

      <section class="curriculum-topic-section">
        <h2>Official curriculum references</h2>
        <ul class="curriculum-source-list">{source_items}</ul>
      </section>
    </div>

    <aside class="curriculum-sidebar">
      <section class="curriculum-panel">
        <h2>Quick links</h2>
        <div class="curriculum-link-row">
          <a class="curriculum-button primary" href="#topic-guide">Topic guide</a>
          <a class="curriculum-button" href="#teacher-slide">Teacher slide</a>
          <a class="curriculum-button" href="/{h(unit['yearFolder'])}/curriculum/">All {h(label)} units</a>
        </div>
      </section>
      <section class="curriculum-panel">
        <h2>Learning path</h2>
        <div class="curriculum-link-row">{''.join(vertical_links)}</div>
      </section>
    </aside>
  </main>

  <p class="curriculum-footer-meta">Feedback: <a href="mailto:{SUPPORT_EMAIL}">{SUPPORT_EMAIL}</a>.</p>
</div>
<script>
  window.skillrPageMeta = {{
    pageType: "topic guide",
    year: "{h(label)}",
    subject: "{h(subject_label)}",
    curriculumCode: "{h(unit['code'])}",
    title: "{h(unit['title'])}",
    supportEmail: "{SUPPORT_EMAIL}"
  }};
  window.skillrAccess = {{
    product: "curriculum-topic-guide",
    accessLevel: "free",
    requiresLogin: false,
    requiresPayment: false
  }};
</script>
<script src="/assets/access.js?v=1"></script>
<script src="/assets/report-issue.js?v=1"></script>
<script src="/pwa-register.js"></script>
</body>
</html>
"""
    output = ctx.repo_root / unit["yearFolder"] / unit["subjectSlug"] / unit["unitSlug"] / "index.html"
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(html_text, encoding="utf-8")


def build_pages(ctx: BuildContext, level: str) -> None:
    year_units = [unit for unit in ctx.units if unit["level"] == level]
    # Year gateways and subject hubs are generated site-wide by
    # scripts/build_year_curriculum_hubs.py so this per-year build cannot
    # accidentally restore the retired long pilot landing page.
    for unit in year_units:
        generate_topic_page(ctx, unit, year_units)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    source = parser.add_mutually_exclusive_group(required=True)
    source.add_argument("--workbook", type=Path, help="Path to curriculum-workbook .xlsx")
    source.add_argument("--from-manifest", action="store_true", help="Regenerate pages from the existing manifest")
    parser.add_argument("--repo-root", default=Path(__file__).resolve().parents[1], type=Path)
    parser.add_argument("--year", default="Year 1", help='Level to generate, or "all"')
    parser.add_argument("--min-year", default=0, type=int, choices=range(0, 11), help="Lowest year to update")
    parser.add_argument("--manifest", default=None, type=Path, help="Manifest input/output path")
    parser.add_argument("--refresh-titles", action="store_true", help="Refresh display titles while preserving URLs")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    repo_root = args.repo_root.resolve()
    manifest_path = (args.manifest or repo_root / "data" / "curriculum-units.json").resolve()

    if args.workbook:
        workbook = args.workbook.resolve()
        units = build_manifest(workbook, repo_root)
        payload = {
            "generatedBy": "scripts/build_curriculum_pages.py",
            "sourceWorkbook": workbook.name,
            "siteOrigin": SITE_ORIGIN,
            "analytics": {"googleAnalyticsId": GA_ID, "adsenseClient": ADSENSE_CLIENT},
            "questionBankStandard": {
                "practiceQuestionsPerUnit": 8,
                "testQuestionsPerUnit": 8,
                "questionsPerAttempt": 8,
                "worksheetQuestions": 8,
                "passingPercent": 75,
                "preReadSeconds": 0,
                "preReadRequired": True,
                "preReadElements": ["learning summary", "key vocabulary", "common trap"],
                "complexTasks": "Prefer printable PDF/worksheet tasks for multi-step or hard-to-model interactions.",
            },
            "units": units,
        }
        write_json(manifest_path, payload)
    else:
        payload = json.loads(manifest_path.read_text(encoding="utf-8"))
        units = payload["units"]
        workbook = Path(payload.get("sourceWorkbook") or "curriculum-units.json")

    for unit in units:
        if unit["yearNumber"] >= args.min_year:
            unit["title"] = title_from_description(unit["code"], unit["description"])

    if args.refresh_titles:
        payload["units"] = units
        write_json(manifest_path, payload)

    if args.year.lower() == "all":
        selected_levels = [level for level in LEVELS if level_number(level) >= args.min_year]
    else:
        if args.year not in LEVELS:
            raise ValueError(f"Unsupported level: {args.year}")
        if level_number(args.year) < args.min_year:
            raise ValueError(f"{args.year} is below --min-year {args.min_year}")
        selected_levels = [args.year]

    ctx = BuildContext(repo_root=repo_root, workbook=workbook, manifest_path=manifest_path, units=units)
    for level in selected_levels:
        build_pages(ctx, level)

    print(json.dumps({
        "units": len(units),
        "generatedLevels": selected_levels,
        "generatedUnits": sum(1 for unit in units if unit["level"] in selected_levels),
        "manifest": str(manifest_path),
    }, indent=2))


if __name__ == "__main__":
    main()
